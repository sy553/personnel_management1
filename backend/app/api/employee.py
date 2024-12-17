import os
from flask import Blueprint, request, jsonify, current_app, send_file, send_from_directory
from werkzeug.utils import secure_filename
from app.models.employee import Employee, EducationHistory, WorkHistory, PositionChangeHistory, ContractAttachment
from app.models.department import Department
from app.models.position import Position
from app import db
import uuid
from sqlalchemy.orm import joinedload
from sqlalchemy import or_, func, text
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask_cors import CORS
import mimetypes

bp = Blueprint('employee', __name__, url_prefix='/api/employees')

# 配置CORS
CORS(bp, resources={
    r"/*": {
        "origins": ["http://localhost:3000"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization", "X-Requested-With", "Accept"],
        "expose_headers": ["Content-Type", "Authorization"],
        "allow_credentials": True
    }
}, supports_credentials=True)

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', 'uploads')
PHOTOS_FOLDER = os.path.join(UPLOAD_FOLDER, 'photos')
CONTRACTS_FOLDER = os.path.join(UPLOAD_FOLDER, 'contracts')

for folder in [UPLOAD_FOLDER, PHOTOS_FOLDER, CONTRACTS_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

def generate_unique_filename(original_filename):
    """生成唯一的文件名"""
    ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
    return f"{uuid.uuid4().hex}.{ext}"

@bp.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '没有文件上传'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'msg': '没有选择文件'}), 400
    
    if file:
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}
        if '.' not in file.filename or \
           file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'code': 400, 'msg': '不支持的文件类型'}), 400

        filename = generate_unique_filename(file.filename)
        if file.filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}:
            file_path = os.path.join(PHOTOS_FOLDER, filename)
        else:
            file_path = os.path.join(CONTRACTS_FOLDER, filename)
        file.save(file_path)

        if file.filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}:
            file_url = f'/static/uploads/photos/{filename}'
        else:
            file_url = f'/static/uploads/contracts/{filename}'
        return jsonify({
            'code': 200,
            'msg': '上传成功',
            'data': {'file_url': file_url}
        })

    return jsonify({'code': 400, 'msg': '上传失败'}), 400

@bp.route('/upload/photo/<int:employee_id>', methods=['POST', 'OPTIONS'])
def upload_photo(employee_id):
    """上传员工照片"""
    if request.method == 'OPTIONS':
        return '', 200
        
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            current_app.logger.error(f'No file part in the request for employee {employee_id}')
            return jsonify({'code': 400, 'msg': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            current_app.logger.error(f'No selected file for employee {employee_id}')
            return jsonify({'code': 400, 'msg': '没有选择文件'}), 400
            
        # 验证文件类型
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
        if '.' not in file.filename or \
           file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            current_app.logger.error(f'Invalid file type for employee {employee_id}: {file.filename}')
            return jsonify({'code': 400, 'msg': '不支持的文件类型'}), 400
            
        # 验证员工是否存在
        employee = Employee.query.get(employee_id)
        if not employee:
            current_app.logger.error(f'Employee not found: {employee_id}')
            return jsonify({'code': 404, 'msg': '员工不存在'}), 404
        
        # 生成安全的文件名
        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{employee.employee_id}_{timestamp}.{original_filename.rsplit('.', 1)[1].lower()}"
        
        # 确保上传目录存在
        photos_folder = os.path.join(current_app.config['BASE_DIR'], 'static', 'uploads', 'photos')
        os.makedirs(photos_folder, exist_ok=True)
        
        # 保存文件
        file_path = os.path.join(photos_folder, filename)
        file.save(file_path)
        
        # 更新员工照片URL
        photo_url = f'/static/uploads/photos/{filename}'
        employee.photo_url = photo_url
        db.session.commit()
        
        current_app.logger.info(f'Successfully uploaded photo for employee {employee_id}: {photo_url}')
        
        return jsonify({
            'code': 200,
            'msg': '上传成功',
            'data': {
                'photo_url': photo_url
            }
        })
        
    except Exception as e:
        current_app.logger.error(f'Error uploading photo for employee {employee_id}: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'msg': '上传失败'}), 500

@bp.route('/upload/contract/<int:employee_id>', methods=['POST', 'OPTIONS'])
def upload_contract(employee_id):
    """上传合同文件"""
    if request.method == 'OPTIONS':
        return '', 200
    
    try:
        if 'file' not in request.files:
            return jsonify({'code': 400, 'msg': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'code': 400, 'msg': '没有选择文件'}), 400
        
        # 验证文件类型
        allowed_extensions = {'pdf', 'doc', 'docx'}
        if '.' not in file.filename or \
           file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'code': 400, 'msg': '不支持的文件类型'}), 400
        
        # 验证员工是否存在
        employee = Employee.query.get(employee_id)
        if not employee:
            return jsonify({'code': 404, 'msg': '员工不存在'}), 404
        
        # 生成安全的文件名
        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{employee_id}_{timestamp}_{original_filename}"
        file_type = filename.rsplit('.', 1)[1].lower()
        
        # 确保目录存在
        os.makedirs(CONTRACTS_FOLDER, exist_ok=True)
        
        # 保存文件
        file_path = os.path.join(CONTRACTS_FOLDER, filename)
        file.save(file_path)
        
        # 构建文件URL
        file_url = f'/static/uploads/contracts/{filename}'
        
        # 保存到数据库
        try:
            contract = ContractAttachment(
                employee_id=employee_id,
                file_name=original_filename,
                file_url=file_url,
                file_type=file_type,
                upload_time=datetime.now()
            )
            db.session.add(contract)
            db.session.commit()
            
            print(f"Contract saved successfully: {contract.to_dict()}")
            
            return jsonify({
                'code': 200,
                'msg': '上传成功',
                'data': contract.to_dict()
            })
            
        except Exception as e:
            # 如果数据库保存失败，删除已上传的文件
            if os.path.exists(file_path):
                os.remove(file_path)
            print(f"Database error: {str(e)}")
            raise e
        
    except Exception as e:
        db.session.rollback()
        print(f"Error uploading contract: {str(e)}")
        return jsonify({'code': 500, 'msg': f'上传失败: {str(e)}'}), 500

@bp.route('/<int:employee_id>/contracts', methods=['GET'])
def get_employee_contracts(employee_id):
    """获取员工的合同列表"""
    try:
        contracts = ContractAttachment.query.filter_by(employee_id=employee_id).order_by(ContractAttachment.upload_time.desc()).all()
        return jsonify({
            'code': 200,
            'msg': '获取成功',
            'data': [contract.to_dict() for contract in contracts]
        })
    except Exception as e:
        print(f"Error getting contracts: {str(e)}")
        return jsonify({'code': 500, 'msg': '获取合同列表失败'}), 500

@bp.route('/contracts/<filename>/preview', methods=['GET'])
def preview_contract(filename):
    """预览合同文件"""
    try:
        # 验证文件是否存在
        contract = ContractAttachment.query.filter_by(file_name=filename).first()
        if not contract:
            return jsonify({'code': 404, 'msg': '文件不存在'}), 404
        
        file_path = os.path.join(CONTRACTS_FOLDER, f"{contract.employee_id}_{contract.upload_time.strftime('%Y%m%d_%H%M%S')}_{filename}")
        if not os.path.exists(file_path):
            return jsonify({'code': 404, 'msg': '文件不存在'}), 404
        
        # 获取文件的MIME类型
        mime_type = mimetypes.guess_type(file_path)[0]
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        return send_file(
            file_path,
            mimetype=mime_type,
            as_attachment=False,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error previewing contract: {str(e)}")
        return jsonify({'code': 500, 'msg': '预览文件失败'}), 500

@bp.route('/contracts/<filename>/download', methods=['GET'])
def download_contract(filename):
    """下载合同文件"""
    try:
        # 验证文件是否存在
        contract = ContractAttachment.query.filter_by(file_name=filename).first()
        if not contract:
            return jsonify({'code': 404, 'msg': '文件不存在'}), 404
        
        file_path = os.path.join(CONTRACTS_FOLDER, f"{contract.employee_id}_{contract.upload_time.strftime('%Y%m%d_%H%M%S')}_{filename}")
        if not os.path.exists(file_path):
            return jsonify({'code': 404, 'msg': '文件不存在'}), 404
        
        # 获取文件的MIME类型
        mime_type = mimetypes.guess_type(file_path)[0]
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        return send_file(
            file_path,
            mimetype=mime_type,
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error downloading contract: {str(e)}")
        return jsonify({'code': 500, 'msg': '下载文件失败'}), 500

@bp.route('', methods=['GET'])
def get_employees():
    """获取员工列表，支持搜索和分页"""
    try:
        # 获取查询参数，添加类型转换和验证
        try:
            page = max(1, int(request.args.get('page', 1)))
            per_page = max(1, min(100, int(request.args.get('per_page', 10))))
        except ValueError:
            page = 1
            per_page = 10
            
        search = request.args.get('search', '').strip()
        department_id = request.args.get('department_id')
        position_id = request.args.get('position_id')
        education = request.args.get('education')
        employment_status = request.args.get('employment_status')
        
        # 打印请求参数
        print(f"\n请求参数:")
        print(f"page: {page}")
        print(f"per_page: {per_page}")
        print(f"search: {search}")
        print(f"department_id: {department_id}")
        print(f"position_id: {position_id}")
        print(f"education: {education}")
        print(f"employment_status: {employment_status}")
        
        # 构建基础查询
        query = Employee.query.options(
            joinedload(Employee.department),
            joinedload(Employee.position)
        )

        # 添加过滤条件
        if search:
            query = query.filter(
                or_(
                    Employee.name.like(f'%{search}%'),
                    Employee.employee_id.like(f'%{search}%'),
                    Employee.phone.like(f'%{search}%'),
                    Employee.id_card.like(f'%{search}%')
                )
            )
            
        # 处理department_id参数
        if department_id and department_id != 'undefined' and department_id != 'null':
            try:
                dept_id = int(department_id)
                query = query.filter(Employee.department_id == dept_id)
            except ValueError:
                print(f"Invalid department_id value: {department_id}")
            
        # 处理position_id参数
        if position_id and position_id != 'undefined' and position_id != 'null':
            try:
                pos_id = int(position_id)
                query = query.filter(Employee.position_id == pos_id)
            except ValueError:
                print(f"Invalid position_id value: {position_id}")
            
        # 处理education参数
        if education and education != 'undefined' and education != 'null':
            query = query.filter(Employee.education == education)

        # 添加在职状态筛选
        valid_statuses = ['active', 'resigned', 'suspended']
        if employment_status and employment_status != 'undefined' and employment_status != 'null':
            if employment_status in valid_statuses:
                print(f"\n正在筛选状态为 {employment_status} 的员工")
                query = query.filter(Employee.employment_status == employment_status)
            else:
                print(f"\n无效的employment_status值: {employment_status}")
        
        # 打印SQL查询语句
        print("\nSQL查询:")
        print(str(query))

        # 执行分页查询
        pagination = query.paginate(page=page, per_page=per_page)
        
        # 打印查询结果
        print(f"\n查询结果: 总数 {pagination.total}")
        for item in pagination.items:
            print(f"ID: {item.id}, Name: {item.name}, Status: {item.employment_status}")
        
        return jsonify({
            'code': 200,
            'msg': '获取员工列表成功',
            'data': {
                'total': pagination.total,
                'items': [item.to_dict() for item in pagination.items]
            }
        })
        
    except Exception as e:
        print(f'\n获取员工列表失败: {str(e)}')
        return jsonify({
            'code': 500,
            'msg': f'获取员工列表失败: {str(e)}'
        })

@bp.route('/<int:id>', methods=['GET'])
def get_employee(id):
    try:
        employee = Employee.query.options(
            joinedload(Employee.department),
            joinedload(Employee.position)
        ).get(id)
        if not employee:
            return jsonify({
                'code': 404,
                'msg': '员工不存在'
            })
        return jsonify({
            'code': 200,
            'data': employee.to_dict(),
            'msg': '获取员工信息成功'
        })
    except Exception as e:
        return jsonify({
            'code': 500,
            'msg': f'获取员工信息失败: {str(e)}'
        })

@bp.route('', methods=['POST'])
def create_employee():
    try:
        data = request.get_json()
        print('接收到的数据:', data)
        employee = Employee(
            employee_id=data['employee_id'],
            name=data['name'],
            gender=data.get('gender'),
            education=data.get('education'),
            birth_date=data.get('birth_date'),
            id_card=data.get('id_card'),
            phone=data.get('phone'),
            address=data.get('address'),
            department_id=data.get('department_id'),
            position_id=data.get('position_id'),
            hire_date=data.get('hire_date'),
            employment_status=data.get('employment_status', 'active')
        )
        db.session.add(employee)
        db.session.commit()
        return jsonify({
            'code': 200,
            'data': employee.to_dict(),
            'msg': '创建员工成功'
        })
    except Exception as e:
        db.session.rollback()
        print('创建员工失败:', str(e))
        return jsonify({
            'code': 500,
            'msg': f'创建员工失败: {str(e)}'
        })

@bp.route('/<int:id>', methods=['PUT'])
def update_employee(id):
    try:
        employee = Employee.query.get(id)
        if not employee:
            return jsonify({
                'code': 404,
                'msg': '员工不存在'
            })
        
        data = request.get_json()
        if not data:
            return jsonify({
                'code': 400,
                'msg': '无效的请求数据'
            })
            
        # 记录旧状态
        old_status = employee.employment_status
        
        # 处理部门关联
        if 'department_id' in data:
            department = Department.query.get(data['department_id'])
            if department:
                employee.department = department
                
        # 处理职位关联
        if 'position_id' in data:
            position = Position.query.get(data['position_id'])
            if position:
                employee.position = position
        
        # 更新其他字段
        for key, value in data.items():
            if hasattr(employee, key) and key not in ['department', 'position', 'department_id', 'position_id']:
                if value is not None:  # 只更新非空值
                    setattr(employee, key, value)
        
        # 记录状态变更
        if 'employment_status' in data and old_status != data['employment_status']:
            current_app.logger.info(f'员工状态变更: {old_status} -> {data["employment_status"]}')
        
        try:
            db.session.commit()
            return jsonify({
                'code': 200,
                'data': employee.to_dict(),
                'msg': '更新员工信息成功'
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'数据库更新失败: {str(e)}')
            return jsonify({
                'code': 500,
                'msg': f'数据库更新失败: {str(e)}'
            })
            
    except Exception as e:
        current_app.logger.error(f'更新员工信息失败: {str(e)}')
        return jsonify({
            'code': 500,
            'msg': f'更新员工信息失败: {str(e)}'
        })

@bp.route('/<int:id>', methods=['DELETE'])
def delete_employee(id):
    try:
        employee = Employee.query.get(id)
        if not employee:
            return jsonify({
                'code': 404,
                'msg': '员工不存在'
            })
        
        db.session.delete(employee)
        db.session.commit()
        return jsonify({
            'code': 200,
            'msg': '删除员工成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'code': 500,
            'msg': f'删除员工失败: {str(e)}'
        })

@bp.route('/<int:employee_id>/education', methods=['GET'])
def get_education_history(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        education_history = employee.education_history.all()
        return jsonify({
            'code': 200,
            'data': [history.to_dict() for history in education_history]
        })
    except Exception as e:
        print('获取教育经历失败:', str(e))
        return jsonify({'code': 500, 'msg': '获取教育经历失败'})

@bp.route('/<int:employee_id>/education', methods=['POST'])
def add_education_history(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        data = request.get_json()
        
        required_fields = ['school', 'degree', 'start_date']
        for field in required_fields:
            if field not in data:
                return jsonify({'code': 400, 'msg': f'缺少必填字段: {field}'})
        
        # 处理日期格式 YYYY-MM -> YYYY-MM-01
        def format_date(date_str):
            if not date_str:
                return None
            try:
                # 尝试解析YYYY-MM格式
                date_obj = datetime.strptime(date_str, '%Y-%m')
                return date_obj.date()
            except ValueError:
                try:
                    # 如果失败，尝试解析YYYY-MM-DD格式
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    return date_obj.date()
                except ValueError as e:
                    print(f"日期格式错误: {str(e)}")
                    raise ValueError(f"无效的日期格式: {date_str}")
        
        education = EducationHistory(
            employee_id=employee_id,
            school=data['school'],
            major=data.get('major'),
            degree=data['degree'],
            start_date=format_date(data['start_date']),
            end_date=format_date(data.get('end_date'))
        )
        
        db.session.add(education)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'data': education.to_dict(),
            'msg': '添加教育经历成功'
        })
    except ValueError as e:
        print('添加教育经历失败:', str(e))
        return jsonify({
            'code': 400,
            'msg': str(e)
        })
    except Exception as e:
        db.session.rollback()
        print('添加教育经历失败:', str(e))
        return jsonify({
            'code': 500,
            'msg': f'添加教育经历失败: {str(e)}'
        })

@bp.route('/<int:employee_id>/work', methods=['GET'])
def get_work_history(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        work_history = employee.work_history.all()
        return jsonify({
            'code': 200,
            'data': [history.to_dict() for history in work_history]
        })
    except Exception as e:
        print('获取工作经历失败:', str(e))
        return jsonify({'code': 500, 'msg': '获取工作经历失败'})

@bp.route('/<int:employee_id>/work', methods=['POST'])
def add_work_history(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        data = request.get_json()
        
        required_fields = ['company', 'position', 'start_date']
        for field in required_fields:
            if field not in data:
                return jsonify({'code': 400, 'msg': f'缺少必填字段: {field}'})
        
        # 处理日期格式 YYYY-MM -> YYYY-MM-01
        def format_date(date_str):
            if not date_str:
                return None
            try:
                # 尝试解析YYYY-MM格式
                date_obj = datetime.strptime(date_str, '%Y-%m')
                return date_obj.date()
            except ValueError:
                try:
                    # 如果失败，尝试解析YYYY-MM-DD格式
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    return date_obj.date()
                except ValueError as e:
                    print(f"日期格式错误: {str(e)}")
                    raise ValueError(f"无效的日期格式: {date_str}")
        
        work = WorkHistory(
            employee_id=employee_id,
            company=data['company'],
            position=data['position'],
            start_date=format_date(data['start_date']),
            end_date=format_date(data.get('end_date'))
        )
        
        db.session.add(work)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'data': work.to_dict(),
            'msg': '添加工作经历成功'
        })
    except ValueError as e:
        print('添加工作经历失败:', str(e))
        return jsonify({
            'code': 400,
            'msg': str(e)
        })
    except Exception as e:
        db.session.rollback()
        print('添加工作经历失败:', str(e))
        return jsonify({
            'code': 500,
            'msg': f'添加工作经历失败: {str(e)}'
        })

@bp.route('/<int:employee_id>/position-changes', methods=['GET'])
def get_position_changes(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        position_changes = employee.position_changes.all()
        return jsonify({
            'code': 200,
            'data': [change.to_dict() for change in position_changes]
        })
    except Exception as e:
        print('获取调岗记录失败:', str(e))
        return jsonify({'code': 500, 'msg': '获取调岗记录失败'})

@bp.route('/<int:employee_id>/position-changes', methods=['POST'])
def add_position_change(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        data = request.get_json()
        
        required_fields = ['new_department_id', 'new_position_id', 'change_date']
        for field in required_fields:
            if field not in data:
                return jsonify({'code': 400, 'msg': f'缺少必填字段: {field}'})
        
        position_change = PositionChangeHistory(
            employee_id=employee_id,
            old_department_id=employee.department_id,
            new_department_id=data['new_department_id'],
            old_position_id=employee.position_id,
            new_position_id=data['new_position_id'],
            change_date=datetime.strptime(data['change_date'], '%Y-%m-%d').date(),
            change_reason=data.get('change_reason')
        )
        
        employee.department_id = data['new_department_id']
        employee.position_id = data['new_position_id']
        
        db.session.add(position_change)
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'msg': '添加成功',
            'data': position_change.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        print('添加调岗记录失败:', str(e))
        return jsonify({'code': 500, 'msg': '添加调岗记录失败'})

@bp.route('/export', methods=['GET'])
def export_employees():
    """导出员工信息"""
    try:
        # 获取状态参数
        status = request.args.get('status', 'active')
        
        # 状态映射
        status_map = {
            'active': '在职',
            'resigned': '离职',
            'suspended': '休假'  # 修改这里
        }
        
        # 查询指定状态的员工
        query = Employee.query.options(
            joinedload(Employee.department),
            joinedload(Employee.position)
        )
        
        if status:
            query = query.filter(Employee.employment_status == status)
            
        employees = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息"
        
        headers = [
            '员工编号', '姓名', '性别', '出生日期', '身份证号',
            '联系电话', '住址', '学历', '部门', '职位',
            '入职日期', '在职状态'
        ]
        
        # 设置表头样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp.employee_id)
            ws.cell(row=row, column=2, value=emp.name)
            ws.cell(row=row, column=3, value=emp.gender)
            ws.cell(row=row, column=4, value=emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '')
            ws.cell(row=row, column=5, value=emp.id_card)
            ws.cell(row=row, column=6, value=emp.phone)
            ws.cell(row=row, column=7, value=emp.address)
            ws.cell(row=row, column=8, value=emp.education)
            ws.cell(row=row, column=9, value=emp.department.name if emp.department else '')
            ws.cell(row=row, column=10, value=emp.position.name if emp.position else '')
            ws.cell(row=row, column=11, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '')
            # 转换状态为中文
            ws.cell(row=row, column=12, value=status_map.get(emp.employment_status, emp.employment_status))
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # 保存到内存
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 生成文件名
        status_text = status_map.get(status, '全部')
        filename = f'employees_{status_text}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f'导出失败: {str(e)}')
        return jsonify({'code': 500, 'msg': f'导出失败: {str(e)}'}), 500

@bp.route('/import', methods=['POST'])
def import_employees():
    """导入员工数据"""
    try:
        current_app.logger.info('开始处理员工导入请求')
        
        if 'file' not in request.files:
            current_app.logger.error('没有上传文件')
            return jsonify({'code': 400, 'msg': '没有上传文件'}), 400
            
        file = request.files['file']
        if file.filename == '':
            current_app.logger.error('没有选择文件')
            return jsonify({'code': 400, 'msg': '没有选择文件'}), 400
            
        if not file.filename.endswith('.xlsx'):
            current_app.logger.error(f'文件格式错误: {file.filename}')
            return jsonify({'code': 400, 'msg': '请上传Excel文件(.xlsx)'}), 400
            
        try:
            current_app.logger.info(f'开始读取Excel文件: {file.filename}')
            df = pd.read_excel(file)
            current_app.logger.info(f'Excel文件读取成功，共{len(df)}行数据')
        except Exception as e:
            current_app.logger.error(f'读取Excel文件失败: {str(e)}')
            return jsonify({'code': 400, 'msg': '无法读取Excel文件，请确保文件格式正确'}), 400
        
        # 验证必填列是否存在
        required_columns = {
            '员工编号*': '员工编号',
            '姓名*': '姓名',
            '部门*': '部门',
            '职位*': '职位'
        }
        current_app.logger.info(f'开始验证必填列，当前列: {list(df.columns)}')
        
        # 重命名带星号的列名
        df.rename(columns={
            '员工编号*': '员工编号',
            '姓名*': '姓名',
            '部门*': '部门',
            '职位*': '职位',
        }, inplace=True)
        
        # 检查必填列是否存在
        missing_columns = [col for col in required_columns.values() if col not in df.columns]
        if missing_columns:
            current_app.logger.error(f'缺少必填列: {missing_columns}')
            return jsonify({
                'code': 400,
                'msg': f'Excel文件缺少必填列: {", ".join(missing_columns)}'
            }), 400
        
        # 获取部门和职位映射
        departments = {dept.name: dept.id for dept in Department.query.all()}
        positions = {pos.name: pos.id for pos in Position.query.all()}
        current_app.logger.info(f'获取到部门列表: {list(departments.keys())}')
        current_app.logger.info(f'获取到职位列表: {list(positions.keys())}')
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 处理每一行数据
        for index, row in df.iterrows():
            try:
                current_app.logger.info(f'处理第{index+2}行数据')
                
                # 验证必填字段
                missing_fields = []
                for field in required_columns.values():
                    value = str(row[field]).strip() if not pd.isna(row[field]) else ''
                    if value == '':
                        missing_fields.append(field)
                
                if missing_fields:
                    error_count += 1
                    error_msg = f'第{index+2}行: {", ".join(missing_fields)}为必填项'
                    errors.append(error_msg)
                    current_app.logger.error(error_msg)
                    continue
                
                # 验证部门存在
                department = str(row['部门']).strip()
                if department not in departments:
                    error_count += 1
                    error_msg = f'第{index+2}行: 部门 {department} 不存在'
                    errors.append(error_msg)
                    current_app.logger.error(error_msg)
                    continue
                
                # 验证职位存在
                position = str(row['职位']).strip()
                if position not in positions:
                    error_count += 1
                    error_msg = f'第{index+2}行: 职位 {position} 不存在'
                    errors.append(error_msg)
                    current_app.logger.error(error_msg)
                    continue
                
                # 处理员工编号
                employee_id = str(row['员工编号']).strip()
                current_app.logger.info(f'处理员工编号: {employee_id}')
                
                # 查找或创建员工
                employee = Employee.query.filter_by(employee_id=employee_id).first()
                if not employee:
                    employee = Employee()
                    current_app.logger.info(f'创建新员工: {employee_id}')
                else:
                    current_app.logger.info(f'更新现有员工: {employee_id}')
                
                # 设置员工信息
                employee.employee_id = employee_id
                employee.name = str(row['姓名']).strip()
                employee.gender = str(row['性别']).strip() if not pd.isna(row['性别']) else None
                employee.birth_date = pd.to_datetime(row['出生日期']).date() if not pd.isna(row['出生日期']) else None
                employee.id_card = str(row['身份证号']).strip() if not pd.isna(row['身份证号']) else None
                employee.phone = str(row['联系电话']).strip() if not pd.isna(row['联系电话']) else None
                employee.address = str(row['住址']).strip() if not pd.isna(row['住址']) else None
                employee.education = str(row['学历']).strip() if not pd.isna(row['学历']) else None
                employee.department_id = departments[department]
                employee.position_id = positions[position]
                employee.hire_date = pd.to_datetime(row['入职日期']).date() if not pd.isna(row['入职日期']) else None
                
                # 状态值映射
                status_mapping = {
                    '在职': 'active',
                    '休假': 'suspended',  # 将"停职"改为"休假"
                    '离职': 'resigned',
                    # 默认值
                    '': 'active'
                }
                
                # 获取状态值并进行映射转换
                status = str(row['在职状态']).strip() if not pd.isna(row['在职状态']) else ''
                employee.employment_status = status_mapping.get(status, 'active')
                
                db.session.add(employee)
                success_count += 1
                current_app.logger.info(f'成功处理第{index+2}行数据')
                
            except Exception as e:
                error_count += 1
                error_msg = f'第{index+2}行: {str(e)}'
                errors.append(error_msg)
                current_app.logger.error(f'处理第{index+2}行时出错: {str(e)}')
        
        # 提交事务
        try:
            current_app.logger.info('开始提交事务')
            db.session.commit()
            current_app.logger.info('事务提交成功')
        except Exception as e:
            db.session.rollback()
            error_msg = f'提交事务失败: {str(e)}'
            current_app.logger.error(error_msg)
            return jsonify({'code': 500, 'msg': f'保存数据失败: {str(e)}'}), 500
        
        result = {
            'code': 200,
            'msg': '导入完成',
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
        }
        current_app.logger.info(f'导入完成: {result}')
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        error_msg = f'导入失败: {str(e)}'
        current_app.logger.error(error_msg)
        return jsonify({'code': 500, 'msg': error_msg}), 500

@bp.route('/import-template', methods=['GET'])
def get_import_template():
    """获取员工导入模板"""
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息模板"
        
        # 设置表头
        headers = [
            '员工编号*', '姓名*', '性别', '出生日期', '身份证号',
            '联系电话', '住址', '学历', '部门*', '职位*',
            '入职日期', '在职状态'
        ]
        
        # 写入表头并设置样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 添加示例数据，修改示例中的在职状态为"在职"
        example = [
            'EMP001', '张三', '男', '1990-01-01', '110101199001011234',
            '13800138000', '北京市朝阳区', '本科', '技术部', '工程师',
            '2020-01-01', '在职'  # 确保示例数据使用"在职"
        ]
        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # 保存到内存
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 设置响应头
        filename = 'employee_import_template.xlsx'
        response = send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'获取模板失败: {str(e)}')
        return jsonify({
            'code': 500,
            'msg': f'获取模板失败: {str(e)}'
        }), 500

@bp.route('/stats', methods=['GET'])
def get_employee_stats():
    """获取员工统计信息"""
    try:
        # 获取员工状态统计
        status_stats = db.session.query(
            Employee.employment_status,
            db.func.count(Employee.id)
        ).group_by(Employee.employment_status).all()
        
        # 获取员工类型统计
        type_stats = db.session.query(
            Employee.employee_type,
            db.func.count(Employee.id)
        ).group_by(Employee.employee_type).all()
        
        # 格式化状态统计结果
        status_count = {
            'active': 0,
            'suspended': 0,
            'resigned': 0,
            'total': 0
        }
        for status, count in status_stats:
            if status in status_count:
                status_count[status] = count
                status_count['total'] += count
                
        # 格式化类型统计结果
        type_count = {
            'intern': 0,
            'probation': 0,
            'regular': 0
        }
        for emp_type, count in type_stats:
            if emp_type in type_count:
                type_count[emp_type] = count
        
        return jsonify({
            'code': 200,
            'data': {
                'status': status_count,
                'type': type_count
            },
            'msg': '获取员工统计信息成功'
        })
        
    except Exception as e:
        current_app.logger.error(f"获取员工统计信息失败: {str(e)}")
        return jsonify({
            'code': 500,
            'msg': f'获取员工统计信息失败: {str(e)}'
        })

@bp.route('/check_photo/<employee_id>', methods=['GET'])
def check_photo(employee_id):
    """检查员工照片是否存在"""
    try:
        employee = Employee.query.get(employee_id)
        if not employee:
            current_app.logger.error(f'Employee not found: {employee_id}')
            return jsonify({'code': 404, 'msg': '员工不存在'}), 404
            
        if not employee.photo_url:
            current_app.logger.info(f'No photo URL for employee {employee_id}')
            return jsonify({
                'code': 404,
                'msg': '员工没有照片',
                'data': {'has_photo': False}
            }), 404
            
        # 获取照片的完整路径
        photo_path = os.path.join(
            current_app.config['BASE_DIR'],
            'app',
            employee.photo_url.lstrip('/')
        )
        
        current_app.logger.info(f'Checking photo path: {photo_path}')
        if os.path.exists(photo_path):
            file_size = os.path.getsize(photo_path)
            return jsonify({
                'code': 200,
                'msg': '照片存在',
                'data': {
                    'has_photo': True,
                    'photo_url': f'/static{employee.photo_url}',
                    'file_size': file_size,
                    'file_path': photo_path
                }
            })
        else:
            current_app.logger.error(f'Photo file not found: {photo_path}')
            # 如果文件不存在，清除数据库中的照片URL
            employee.photo_url = None
            db.session.commit()
            return jsonify({
                'code': 404,
                'msg': '照片文件不存在',
                'data': {'has_photo': False}
            }), 404
            
    except Exception as e:
        current_app.logger.error(f"Error checking photo: {str(e)}")
        return jsonify({
            'code': 500,
            'msg': f'检查照片失败: {str(e)}'
        }), 500
